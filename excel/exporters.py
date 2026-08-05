"""openpyxl builders shared by the inbound/outbound doc exports and the
report screens. Both return an in-memory BytesIO ready to send as an
attachment — callers never write to disk."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

_THIN = Side(style="thin", color="999999")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")


def export_document(title, header_fields, header_data, line_columns, lines):
    """Single-record 'document style' export: header block on top, a
    bordered line-item table below.

    header_fields / line_columns: list of (label, key) pairs.
    header_data: dict.  lines: list of dicts.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Document"

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    for label, key in header_fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=header_data.get(key))
        row += 1

    row += 1
    header_row = row
    for col_idx, (label, _key) in enumerate(line_columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")

    for line in lines:
        row += 1
        for col_idx, (_label, key) in enumerate(line_columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=line.get(key))
            cell.border = BORDER

    for col_idx in range(1, len(line_columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_workbook(sheets):
    """Full-database-style backup: one workbook, one sheet per table.

    sheets: list of (sheet_title, columns, rows). columns: list of
    (label, key) pairs. rows: list of dicts.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for title, columns, rows in sheets:
        ws = wb.create_sheet(title=title[:31])
        for col_idx, (label, _key) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.border = BORDER

        for r_idx, row_data in enumerate(rows, start=2):
            for col_idx, (_label, key) in enumerate(columns, start=1):
                cell = ws.cell(row=r_idx, column=col_idx, value=row_data.get(key))
                cell.border = BORDER

        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_table(title, columns, rows):
    """Plain tabular export for the report/query screens.

    columns: list of (label, key) pairs. rows: list of dicts.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)

    header_row = 3
    for col_idx, (label, _key) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for r_idx, row_data in enumerate(rows, start=header_row + 1):
        for col_idx, (_label, key) in enumerate(columns, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=row_data.get(key))
            cell.border = BORDER

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
